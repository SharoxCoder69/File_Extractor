import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import re

STORES = [
    "TWGSC18 -CEDAR LANE SC",
    "TWGSC20 -EASLEY SC",
    "TWGSC21 -MONITOR SC",
    "TWGSC22 -SHOCKLEY SC",
    "TWGSC31 -LAURENS SC",
    "TWGSC66 -ANDERSON MAIN ST SC",
    "TWGVA24 -HIGH ST VA",
    "TWGVA25 -GEORGE W. VA",
    "TWGVA26 -KECOUGHTAN VA",
    "TWGVA27 -NORFOLK VA",
    "TWGVA28 -ABERDEEN VA",
    "TWGVA40 -VIRGINIA BEACH VA",
    "TWGVA58 -WEST MERCURY BLVD VA",
    "TWGVA59 -BATTLEFIELD BLVD VA",
    "TWGVA60 -GREAT NECK RD VA",
    "TWGVA61 -WARWICK BLVD VA",
    "TWGVA62 - NEWMARKET DR VA",
    "TWGVA63 -J CLYDE MORRIS VA",
]

raw_file = ""
report_file = ""

def select_raw():
    global raw_file
    raw_file = filedialog.askopenfilename(
        title="Select Raw Data File"
    )
    raw_label.config(text=raw_file)

def select_report():
    global report_file
    report_file = filedialog.askopenfilename(
        title="Select EMP Report File"
    )
    report_label.config(text=report_file)

def extract_ntids(text):
    ntids = set()

    for store in STORES:

        if store not in text:
            continue

        start = text.find(store)

        next_store_pos = len(text)

        for s in STORES:
            pos = text.find(s, start + 1)

            if pos != -1 and pos < next_store_pos:
                next_store_pos = pos

        block = text[start:next_store_pos]

        for line in block.splitlines():

            match = re.match(r'^([A-Z]{3}\d{5})', line.strip())

            if match:
                ntids.add(match.group(1))

    return ntids

def process():

    if not raw_file or not report_file:
        messagebox.showerror("Error", "Dono files select karo")
        return

    try:

        with open(raw_file, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()

        ntids = extract_ntids(text)

        wb = load_workbook(report_file)

        for ws in wb.worksheets:

            for row in range(1, ws.max_row + 1):

                ntid = ws.cell(row=row, column=3).value

                if ntid in ntids:

                    for col in range(5, 11):
                        ws.cell(row=row, column=col).value = "Yes"

        output = report_file.replace(".xlsx", "_UPDATED.xlsx")

        wb.save(output)

        messagebox.showinfo(
            "Done",
            f"Report Ready!\n\nSaved:\n{output}"
        )

    except Exception as e:
        messagebox.showerror("Error", str(e))

root = tk.Tk()
root.title("Store Employee YES Generator")
root.geometry("700x300")

tk.Button(
    root,
    text="Select Raw Data",
    command=select_raw,
    width=30
).pack(pady=10)

raw_label = tk.Label(root, text="No Raw Data Selected")
raw_label.pack()

tk.Button(
    root,
    text="Select EMP Report",
    command=select_report,
    width=30
).pack(pady=10)

report_label = tk.Label(root, text="No Report Selected")
report_label.pack()

tk.Button(
    root,
    text="Generate Report",
    command=process,
    width=30,
    height=2
).pack(pady=20)

root.mainloop()
